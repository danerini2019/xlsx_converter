import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import xlrd

def convert_to_xlsx(input_file):
    output_file = os.path.splitext(input_file)[0] + '.xlsx'
    wb = Workbook()
    ws = wb.active

    # Open the .xls file
    xls_workbook = xlrd.open_workbook(input_file)
    xls_sheet = xls_workbook.sheet_by_index(0)

    # Iterate through each row and column in the .xls file and copy the data to the .xlsx file
    for row_index in range(xls_sheet.nrows):
        for col_index in range(xls_sheet.ncols):
            cell_value = xls_sheet.cell_value(row_index, col_index)
            ws[get_column_letter(col_index + 1) + str(row_index + 1)] = cell_value

    # Save the .xlsx file
    wb.save(output_file)

    print(f"File '{input_file}' converted to '{output_file}'")

def main():
    # Specify the directory containing .xls files
    input_directory = r'\\RCK-SVR-FILE.TYLI.COM\TYLin-RCK$\Projects\3010.0101541.000\400_DATA\50_Traffic\Data Collection_May 2024\Imperial Traffic and Data Collection\TMC Data\xls to xlsx converted'

    # Get a list of .xls files in the directory
    xls_files = [f for f in os.listdir(input_directory) if f.endswith('.xls')]

    # Convert each .xls file to .xlsx
    for xls_file in xls_files:
        convert_to_xlsx(os.path.join(input_directory, xls_file))

if __name__ == "__main__":
    main()
